from __future__ import annotations

import argparse
import json
import math
import shutil
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ID = "pnj-compound-company-limited"
FIRESTORE_BASE = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents"
DEFAULT_MONTH_ID = "202605"


@dataclass
class ShareholderRecord:
    uid: str
    shareholder_code: str
    display_name: str
    active: bool


def run_json(command: list[str]) -> Any:
    result = subprocess.run(command, check=True, capture_output=True, text=True, encoding="utf-8")
    return json.loads(result.stdout)


def find_npx() -> str:
    candidates = [
        shutil.which("npx"),
        shutil.which("npx.cmd"),
        r"C:\Program Files\nodejs\npx.cmd",
        r"C:\Program Files\nodejs\npx.exe",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return candidate
    raise RuntimeError("找不到 npx，請先安裝 Node.js 與 firebase-tools。")


def get_access_token() -> str:
    data = run_json([find_npx(), "firebase-tools", "login:list", "--json"])
    accounts = data.get("result") or []
    if not accounts:
        raise RuntimeError("找不到 Firebase 登入資訊，請先登入 firebase-tools。")
    token = accounts[0].get("tokens", {}).get("access_token")
    if not token:
        raise RuntimeError("找不到 Firebase access token。")
    return token


def firestore_request(token: str, method: str, url: str, body: dict[str, Any] | None = None) -> dict[str, Any]:
    data = None
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload) if payload else {}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Firestore API {method} {url} 失敗: {exc.code} {detail}") from exc


def list_shareholders(token: str) -> dict[str, ShareholderRecord]:
    shareholders: dict[str, ShareholderRecord] = {}
    next_token = None
    while True:
        query = {"pageSize": 200}
        if next_token:
            query["pageToken"] = next_token
        url = f"{FIRESTORE_BASE}/shareholders?{urllib.parse.urlencode(query)}"
        data = firestore_request(token, "GET", url)
        for doc in data.get("documents", []):
            fields = parse_firestore_fields(doc.get("fields", {}))
            code = str(fields.get("shareholderCode") or "").upper()
            if not code:
                continue
            shareholders[code] = ShareholderRecord(
                uid=doc["name"].split("/")[-1],
                shareholder_code=code,
                display_name=str(fields.get("displayName") or ""),
                active=fields.get("active", True) is not False,
            )
        next_token = data.get("nextPageToken")
        if not next_token:
            break
    return shareholders


def parse_firestore_fields(fields: dict[str, Any]) -> dict[str, Any]:
    return {key: parse_firestore_value(value) for key, value in fields.items()}


def parse_firestore_value(value: dict[str, Any]) -> Any:
    if "stringValue" in value:
        return value["stringValue"]
    if "integerValue" in value:
        return int(value["integerValue"])
    if "doubleValue" in value:
        return float(value["doubleValue"])
    if "booleanValue" in value:
        return bool(value["booleanValue"])
    if "timestampValue" in value:
        return value["timestampValue"]
    if "nullValue" in value:
        return None
    if "mapValue" in value:
        return parse_firestore_fields(value["mapValue"].get("fields", {}))
    if "arrayValue" in value:
        return [parse_firestore_value(item) for item in value["arrayValue"].get("values", [])]
    return value


def to_firestore_value(value: Any) -> dict[str, Any]:
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int) and not isinstance(value, bool):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError(f"無法寫入非有限數值: {value}")
        if value.is_integer():
            return {"doubleValue": float(value)}
        return {"doubleValue": value}
    if isinstance(value, str):
        return {"stringValue": value}
    if isinstance(value, list):
        values = [to_firestore_value(item) for item in value]
        return {"arrayValue": {"values": values}} if values else {"arrayValue": {}}
    if isinstance(value, dict):
        return {"mapValue": {"fields": {key: to_firestore_value(item) for key, item in value.items()}}}
    raise TypeError(f"不支援的 Firestore 型別: {type(value)!r}")


def to_firestore_fields(data: dict[str, Any]) -> dict[str, Any]:
    return {key: to_firestore_value(value) for key, value in data.items()}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if text.endswith(" 00:00:00"):
        text = text[:10]
    if "/" in text and len(text.split("/")) == 3:
        parts = text.split("/")
        if len(parts[0]) == 4:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return text


def normalize_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value) if isinstance(value, int) or float(value).is_integer() else float(value)
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


def percent_to_display(value: Any) -> float | None:
    number = normalize_number(value)
    if number is None:
        return None
    scaled = float(number) * 100 if abs(float(number)) <= 1 else float(number)
    return round(scaled, 6)


def normalize_code(prefix: Any, seq: Any) -> str:
    joined = f"{clean_text(prefix)}{clean_text(seq)}"
    return joined.replace("-", "").replace(" ", "").upper()


def infer_share_class(raw: str) -> str:
    text = clean_text(raw).upper()
    if text.endswith("CS") or "_CS" in text:
        return "CS"
    if text.endswith("PS") or "_PS" in text:
        return "PS"
    return text


def compose_name(parts: list[str]) -> str:
    values = [clean_text(part) for part in parts if clean_text(part)]
    if len(values) >= 3 and values[-1] == ")" and "(" in values[0]:
        return f"{values[0]}{'、'.join(values[1:-1])})"
    return "".join(values)


def format_statement_label(year: str, month: str) -> str:
    return f"{year} 年 {month} 月"


def contains_any(text: str, needles: list[str]) -> bool:
    lowered = text.lower()
    return any(needle.lower() in lowered for needle in needles)


def find_row_by_first_cell(ws: Any, needles: list[str]) -> int | None:
    for row in range(1, ws.max_row + 1):
        value = clean_text(ws.cell(row, 1).value)
        if contains_any(value, needles):
            return row
    return None


def find_dividend_header_row(ws: Any) -> int | None:
    for row in range(1, ws.max_row + 1):
        year = clean_text(ws.cell(row, 4).value)
        amount = clean_text(ws.cell(row, 5).value)
        method = clean_text(ws.cell(row, 6).value)
        if year == "股利年度" and amount == "金額" and method == "領取方式":
            return row
    return None


def find_contribution_header_row(ws: Any) -> int | None:
    for row in range(1, ws.max_row + 1):
        name = clean_text(ws.cell(row, 1).value)
        amount = clean_text(ws.cell(row, 2).value)
        date_value = clean_text(ws.cell(row, 3).value)
        if "股東名稱" in name and ("投資金額" in amount or "已到位投資金額" in amount) and "投入日期" in date_value:
            return row
    return None


def is_dividend_header_like(year: Any, amount: Any, method: Any) -> bool:
    return clean_text(year) == "股利年度" or clean_text(amount) == "金額" or clean_text(method) == "領取方式"


def is_total_row(name: str, funded_amount: float | int | None, units: float | int | None) -> bool:
    normalized_name = clean_text(name)
    return (
        "合計" in normalized_name or
        "total" in normalized_name.lower() or
        (not normalized_name and funded_amount is not None and units is not None)
    )


def parse_sheet(ws: Any) -> dict[str, Any]:
    name_row = find_row_by_first_cell(ws, ["_Name", "股東姓名"])
    code_row = find_row_by_first_cell(ws, ["_CIN", "戶號"])
    currency_row = find_row_by_first_cell(ws, ["_Currency", "幣別"])
    report_type_row = find_row_by_first_cell(ws, ["_Attributes", "報告性質"])
    end_date_row = find_row_by_first_cell(ws, ["End Date", "結算截止日"])
    nav_row = find_row_by_first_cell(ws, ["淨值", "每單位淨值", "值：", "值:"])
    share_class_row = find_dividend_header_row(ws)
    contribution_header_row = find_contribution_header_row(ws)
    notes_header_row = find_row_by_first_cell(ws, ["Explanatory Notes", "說明事項"])

    if not all([
        name_row,
        code_row,
        currency_row,
        report_type_row,
        end_date_row,
        nav_row,
        share_class_row,
        contribution_header_row,
        notes_header_row,
    ]):
        raise RuntimeError(f"{ws.title} 缺少必要欄位，請檢查模板格式。")

    name = compose_name([ws.cell(name_row, col).value for col in range(2, 7)])
    shareholder_code = normalize_code(ws.cell(code_row, 2).value, ws.cell(code_row, 3).value)
    share_class_raw = clean_text(ws.cell(share_class_row, 2).value)

    dividend_rows: list[dict[str, Any]] = []
    for offset in range(1, 5):
        row = share_class_row + offset
        year_text = clean_text(ws.cell(row, 4).value)
        amount = normalize_number(ws.cell(row, 5).value)
        method = clean_text(ws.cell(row, 6).value)
        if (year_text or amount is not None or method) and not is_dividend_header_like(year_text, amount, method):
            dividend_rows.append({
                "year": year_text,
                "amount": amount if amount is not None else "",
                "method": method,
                "sortOrder": len(dividend_rows) + 1,
            })

    contribution_rows: list[dict[str, Any]] = []
    for row in range(contribution_header_row + 1, notes_header_row):
        raw_name = clean_text(ws.cell(row, 1).value)
        funded_amount = normalize_number(ws.cell(row, 2).value)
        contribution_date = normalize_date(ws.cell(row, 3).value)
        nav_cell = ws.cell(row, 4).value
        buy_nav = normalize_number(ws.cell(row, 5).value)
        units = normalize_number(ws.cell(row, 6).value)
        remark = clean_text(ws.cell(row, 7).value)

        row_values = [raw_name, funded_amount, contribution_date, nav_cell, buy_nav, units, remark]
        if not any(value not in ("", None) for value in row_values):
            continue
        if is_total_row(raw_name, funded_amount, units):
            continue

        nav_date = normalize_date(nav_cell) if isinstance(nav_cell, (datetime, date)) else ""
        nav_label = "" if nav_date else clean_text(nav_cell)

        contribution_rows.append({
            "shareholderName": raw_name or name,
            "fundedAmount": funded_amount if funded_amount is not None else "",
            "contributionDate": contribution_date,
            "navDateLabel": nav_label,
            "navDate": nav_date,
            "buyNavPerUnit": buy_nav if buy_nav is not None else "",
            "unitsAcquired": units if units is not None else "",
            "remark": remark,
            "sortOrder": len(contribution_rows) + 1,
        })

    notes: list[str] = []
    for row in range(notes_header_row + 1, ws.max_row + 1):
        text = clean_text(ws.cell(row, 1).value)
        if text:
            notes.append(text)

    end_date = normalize_date(ws.cell(end_date_row, 2).value)
    year = end_date[:4] if len(end_date) >= 7 else DEFAULT_MONTH_ID[:4]
    month = end_date[5:7] if len(end_date) >= 7 else DEFAULT_MONTH_ID[4:6]

    return {
        "sheetName": ws.title,
        "monthId": f"{year}{month}",
        "payload": {
            "renderMode": "structured-html",
            "contentType": "text/html",
            "templateVersion": 1,
            "year": year,
            "month": month,
            "label": format_statement_label(year, month),
            "reportType": clean_text(ws.cell(report_type_row, 2).value),
            "shareholderCode": shareholder_code,
            "shareholderName": name,
            "shareClass": infer_share_class(share_class_raw),
            "currency": clean_text(ws.cell(currency_row, 2).value) or "USD",
            "endDate": end_date,
            "navPerUnit": normalize_number(ws.cell(nav_row, 2).value),
            "investedAmountUsd": normalize_number(ws.cell(share_class_row + 1, 2).value),
            "unitsHeld": normalize_number(ws.cell(share_class_row + 2, 2).value),
            "ownershipPercent": percent_to_display(ws.cell(share_class_row + 3, 2).value),
            "marketValueUsd": normalize_number(ws.cell(share_class_row + 4, 2).value),
            "gainUsd": normalize_number(ws.cell(share_class_row + 5, 2).value),
            "returnRatePercent": percent_to_display(ws.cell(share_class_row + 6, 2).value),
            "dividendRows": dividend_rows,
            "contributionRows": contribution_rows,
            "notes": notes,
        },
    }


def build_patch_fields(payload: dict[str, Any], status_if_new: str | None) -> dict[str, Any]:
    data = dict(payload)
    data["updatedAt"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    if status_if_new:
        data["status"] = status_if_new
    return data


def get_statement_doc(token: str, uid: str, month_id: str) -> tuple[bool, dict[str, Any]]:
    url = f"{FIRESTORE_BASE}/statements/{uid}/months/{month_id}"
    request = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}", "Accept": "application/json"}, method="GET")
    try:
        with urllib.request.urlopen(request) as response:
            payload = response.read().decode("utf-8")
            return True, json.loads(payload) if payload else {}
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            return False, {}
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"讀取既有 statement 失敗: {exc.code} {detail}") from exc


def patch_statement_doc(token: str, uid: str, month_id: str, fields: dict[str, Any]) -> dict[str, Any]:
    params = [("updateMask.fieldPaths", field_name) for field_name in fields]
    query = urllib.parse.urlencode(params)
    url = f"{FIRESTORE_BASE}/statements/{uid}/months/{month_id}?{query}"
    body = {"fields": to_firestore_fields(fields)}
    return firestore_request(token, "PATCH", url, body)


def preview_rows(parsed_rows: list[dict[str, Any]], shareholders: dict[str, ShareholderRecord]) -> list[dict[str, Any]]:
    rows = []
    for item in parsed_rows:
        payload = item["payload"]
        code = payload["shareholderCode"]
        owner = shareholders.get(code)
        rows.append({
            "sheetName": item["sheetName"],
            "monthId": item["monthId"],
            "shareholderCode": code,
            "shareholderName": payload["shareholderName"],
            "matchedUid": owner.uid if owner else None,
            "matchedDisplayName": owner.display_name if owner else None,
            "shareClass": payload["shareClass"],
            "ownershipPercent": payload["ownershipPercent"],
            "returnRatePercent": payload["returnRatePercent"],
            "dividendRowCount": len(payload["dividendRows"]),
            "contributionRowCount": len(payload["contributionRows"]),
        })
    return rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Excel 檔案路徑")
    parser.add_argument("--month-id", default=DEFAULT_MONTH_ID)
    parser.add_argument("--preview-out", default="tmp/import_preview.json")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.xlsx)
    if not workbook_path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    parsed_rows = [parse_sheet(ws) for ws in wb.worksheets]

    token = get_access_token()
    shareholders = list_shareholders(token)
    preview = preview_rows(parsed_rows, shareholders)

    preview_path = Path(args.preview_out)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    preview_path.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")

    missing_codes = [row["shareholderCode"] for row in preview if not row["matchedUid"]]
    if missing_codes:
        raise RuntimeError(f"下列戶號找不到已建立股東: {', '.join(missing_codes)}")

    if not args.apply:
        print(json.dumps({
            "mode": "preview",
            "count": len(preview),
            "previewFile": str(preview_path),
            "rows": preview,
        }, ensure_ascii=False, indent=2))
        return 0

    results = []
    for item in parsed_rows:
        payload = item["payload"]
        month_id = item["monthId"] or args.month_id
        owner = shareholders[payload["shareholderCode"]]
        exists, existing_doc = get_statement_doc(token, owner.uid, month_id)
        existing_fields = parse_firestore_fields(existing_doc.get("fields", {})) if existing_doc else {}
        existing_status = clean_text(existing_fields.get("status"))
        fields = build_patch_fields(payload, existing_status if exists and existing_status else (None if exists else "draft"))
        patch_statement_doc(token, owner.uid, month_id, fields)
        results.append({
            "shareholderCode": payload["shareholderCode"],
            "shareholderName": payload["shareholderName"],
            "uid": owner.uid,
            "monthId": month_id,
            "existingDoc": exists,
            "statusAction": existing_status or ("created-draft" if not exists else "preserved"),
        })

    print(json.dumps({
        "mode": "apply",
        "count": len(results),
        "previewFile": str(preview_path),
        "results": results,
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8")
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover
        print(str(exc), file=sys.stderr)
        raise
